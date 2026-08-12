"""Session-owned CSV/XLSX uploads exposed to Grafana through signed CSV URLs."""
from __future__ import annotations

import base64
import csv
from _csv import Error as CSVError
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import shutil
import time
import zipfile
from pathlib import Path
from typing import Any, BinaryIO

MAX_UPLOAD_BYTES = 50 * 1024 * 1024
MAX_XLSX_EXPANDED_BYTES = 500 * 1024 * 1024
UPLOAD_RETENTION_SECONDS = 7 * 24 * 60 * 60
UPLOAD_ID_RE = re.compile(r"upload_[a-f0-9]{32}")
UPLOAD_ROOT = Path(os.environ.get("UPLOAD_DATASET_ROOT", Path(__file__).resolve().parent / ".analysis-artifacts" / "uploads"))


def _b64(value: bytes) -> str:
    return base64.urlsafe_b64encode(value).decode().rstrip("=")


def _unb64(value: str) -> bytes:
    return base64.urlsafe_b64decode(value + "=" * (-len(value) % 4))


def _safe_name(value: str) -> str:
    return Path(value or "upload").name[:255]


def _decode_csv(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp950", "big5"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CSV encoding must be UTF-8 or Big5/CP950")


def _normalize_headers(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values, 1):
        base = str(value or "").strip() or f"column_{index}"
        seen[base] = seen.get(base, 0) + 1
        headers.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return headers


def _infer_type(values: list[str]) -> str:
    observed = [value.strip() for value in values if value.strip()][:1000]
    if not observed:
        return "string"
    if all(re.fullmatch(r"-?\d+(?:\.\d+)?", value) and not (len(value) > 1 and value.startswith("0")) for value in observed):
        return "number"
    if all(re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[ T].*)?", value) for value in observed):
        return "date"
    return "string"


def _csv_profile(path: Path) -> tuple[list[dict[str, str]], int]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.reader(handle)
        first = next(reader, None)
        if first is None:
            raise ValueError("uploaded dataset is empty")
        headers = _normalize_headers(first)
        samples: list[list[str]] = [[] for _ in headers]
        rows = 0
        for row in reader:
            rows += 1
            for index in range(min(len(row), len(samples))):
                if len(samples[index]) < 1000:
                    samples[index].append(row[index])
    if rows == 0:
        raise ValueError("uploaded dataset has no data rows")
    return [{"name": name, "type": _infer_type(samples[index])} for index, name in enumerate(headers)], rows


def _write_csv(raw: bytes, destination: Path) -> None:
    text = _decode_csv(raw)
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except CSVError:  # deterministic comma fallback for short/ambiguous samples
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text, newline=""), dialect)
    first = next(reader, None)
    if first is None:
        raise ValueError("uploaded CSV is empty")
    with destination.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(_normalize_headers(first))
        writer.writerows(reader)


def _write_xlsx(raw: bytes, destination: Path, requested_sheet: str | None) -> str:
    from openpyxl import load_workbook
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    try:
        with zipfile.ZipFile(io.BytesIO(raw)) as archive:
            names = [item.filename.lower() for item in archive.infolist()]
            if sum(item.file_size for item in archive.infolist()) > MAX_XLSX_EXPANDED_BYTES:
                raise ValueError("XLSX expanded size exceeds 500 MB")
            if any(name.endswith("vbaproject.bin") or name.startswith("xl/externallinks/") for name in names):
                raise ValueError("XLSX macros and external links are not allowed")
    except zipfile.BadZipFile as exc:
        raise ValueError("XLSX file is invalid") from exc
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True, keep_links=False)
    try:
        sheets = list(workbook.sheetnames)
        if not sheets:
            raise ValueError("uploaded workbook has no sheets")
        if requested_sheet is None and len(sheets) > 1:
            raise ValueError("SHEET_SELECTION_REQUIRED:" + json.dumps(sheets, ensure_ascii=False))
        sheet_name = requested_sheet or sheets[0]
        if sheet_name not in sheets:
            raise ValueError("selected sheet does not exist")
        worksheet = next(sheet for sheet in workbook.worksheets if sheet.title == sheet_name)
        if not isinstance(worksheet, ReadOnlyWorksheet):
            raise ValueError("selected sheet is not a readable worksheet")
        with destination.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                values = ["" if value is None else value for value in row]
                writer.writerow(_normalize_headers(values) if index == 0 else values)
        return sheet_name
    finally:
        workbook.close()


def read_limited(stream: BinaryIO, content_length: int) -> bytes:
    if content_length < 1 or content_length > MAX_UPLOAD_BYTES:
        raise ValueError(f"file must be between 1 byte and {MAX_UPLOAD_BYTES} bytes")
    raw = stream.read(content_length)
    if len(raw) != content_length:
        raise ValueError("upload body length is invalid")
    return raw


def store_upload(*, context: dict[str, str], session_id: str, filename: str, raw: bytes, sheet: str | None = None) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("only .csv and .xlsx uploads are supported")
    upload_id = "upload_" + secrets.token_hex(16)
    root = UPLOAD_ROOT / upload_id
    root.mkdir(parents=True, mode=0o700)
    csv_path = root / "data.csv"
    try:
        if suffix == ".xlsx":
            selected_sheet = _write_xlsx(raw, csv_path, sheet)
        else:
            _write_csv(raw, csv_path)
            selected_sheet = None
        fields, rows = _csv_profile(csv_path)
        metadata = {
            "id": upload_id,
            "org_id": str(context["org_id"]),
            "user_id": str(context["user_id"]),
            "session_id": session_id,
            "filename": _safe_name(filename),
            "sheet": selected_sheet,
            "rows": rows,
            "columns": len(fields),
            "fields": fields,
            "created_at": int(time.time()),
            "expires_at": int(time.time()) + UPLOAD_RETENTION_SECONDS,
        }
        (root / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        return metadata
    except Exception:
        shutil.rmtree(root, ignore_errors=True)
        raise


def _load_metadata(upload_id: str) -> dict[str, Any]:
    if not UPLOAD_ID_RE.fullmatch(upload_id):
        raise PermissionError("invalid upload id")
    try:
        metadata = json.loads((UPLOAD_ROOT / upload_id / "metadata.json").read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise PermissionError("uploaded dataset not found") from exc
    try:
        expired = int(metadata.get("expires_at", 0)) < int(time.time())
    except (TypeError, ValueError) as exc:
        raise PermissionError("uploaded dataset metadata is invalid") from exc
    if expired:
        raise PermissionError("uploaded dataset expired")
    return metadata


def inspect_upload(context: dict[str, str], upload_id: str, session_id: str | None = None) -> dict[str, Any]:
    metadata = _load_metadata(upload_id)
    if metadata.get("org_id") != str(context["org_id"]) or metadata.get("user_id") != str(context["user_id"]):
        raise PermissionError("uploaded dataset owner mismatch")
    if session_id is not None and metadata.get("session_id") != session_id:
        raise PermissionError("uploaded dataset session mismatch")
    return metadata


def delete_upload(context: dict[str, str], upload_id: str, session_id: str) -> None:
    metadata = inspect_upload(context, upload_id, session_id)
    try:
        shutil.rmtree(UPLOAD_ROOT / metadata["id"])
    except FileNotFoundError as exc:
        raise OSError("uploaded dataset no longer exists") from exc


def list_uploads(context: dict[str, str]) -> list[dict[str, Any]]:
    cleanup_expired()
    output = []
    session_id = context.get("session_id")
    for path in UPLOAD_ROOT.glob("upload_*/metadata.json"):
        try:
            metadata = _load_metadata(path.parent.name)
        except PermissionError:
            continue
        if metadata.get("org_id") == str(context["org_id"]) and metadata.get("user_id") == str(context["user_id"]) and (session_id is None or metadata.get("session_id") == session_id):
            output.append(metadata)
    return sorted(output, key=lambda item: item["created_at"], reverse=True)


def sign_csv_url(*, public_base: str, secret: str, metadata: dict[str, Any]) -> str:
    try:
        value = {"id": metadata["id"], "org": metadata["org_id"], "user": metadata["user_id"], "exp": min(int(metadata["expires_at"]), int(time.time()) + 3600)}
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("uploaded dataset metadata is invalid") from exc
    payload = _b64(json.dumps(value, separators=(",", ":"), sort_keys=True).encode())
    signature = _b64(hmac.new(secret.encode(), payload.encode(), hashlib.sha256).digest())
    return public_base.rstrip("/") + "/uploaded-csv/" + payload + "." + signature


def read_signed_csv(token: str, secret: str) -> Path:
    try:
        payload, signature = token.split(".", 1)
        expected = _b64(hmac.new(secret.encode(), payload.encode(), hashlib.sha256).digest())
        if not hmac.compare_digest(signature, expected):
            raise PermissionError("invalid upload signature")
        value = json.loads(_unb64(payload))
        if int(value["exp"]) < int(time.time()):
            raise PermissionError("upload URL expired")
        metadata = inspect_upload({"org_id": str(value["org"]), "user_id": str(value["user"])}, str(value["id"]))
        return UPLOAD_ROOT / metadata["id"] / "data.csv"
    except (KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
        raise PermissionError("invalid upload URL") from exc


def cleanup_expired(now: int | None = None) -> int:
    try:
        effective_now = int(time.time()) if now is None else int(now)
    except (TypeError, ValueError) as exc:
        raise ValueError("cleanup time is invalid") from exc
    removed = 0
    UPLOAD_ROOT.mkdir(parents=True, exist_ok=True, mode=0o700)
    for path in UPLOAD_ROOT.glob("upload_*"):
        try:
            metadata = json.loads((path / "metadata.json").read_text(encoding="utf-8"))
            expired = int(metadata.get("expires_at", 0)) < effective_now
        except (OSError, json.JSONDecodeError, TypeError, ValueError):
            expired = True
        if expired:
            try:
                shutil.rmtree(path)
            except FileNotFoundError:
                pass
            removed += 1
    return removed
