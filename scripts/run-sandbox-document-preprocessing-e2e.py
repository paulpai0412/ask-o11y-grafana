#!/usr/bin/env python3
"""Live XLSX document preprocessing → derived frame + derived dataset E2E."""
from __future__ import annotations

import io
import json
import os
import urllib.error
import urllib.parse
import urllib.request
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

TOKEN = os.environ.get("MCP_SHARED_TOKEN", "")
ORG = os.environ.get("ANALYSIS_SERVICE_ORG_ID", "1")
USER = os.environ.get("ANALYSIS_SERVICE_USER_ID", "ask-o11y")
PLANNER_URL = os.environ.get("DATA_QUERY_PLANNER_MCP_URL", "http://127.0.0.1:8768/mcp")
GRAFANA_QUERY_URL = os.environ.get("GRAFANA_QUERY_MCP_URL", "http://127.0.0.1:8772/mcp")
SANDBOX_URL = os.environ.get("SANDBOX_ANALYSIS_MCP_URL", "http://127.0.0.1:8777/mcp")
UPLOAD_URL = GRAFANA_QUERY_URL.removesuffix("/mcp") + "/uploads"
SESSION_ID = "document-preprocessing-e2e"
HEADERS = {"Authorization": f"Bearer {TOKEN}", "X-Grafana-Org-Id": ORG, "X-Grafana-User": USER, "X-Grafana-Actor-User-Id": USER}


def rpc(url: str, name: str, arguments: dict[str, Any], timeout: int = 300) -> dict[str, Any]:
    request = urllib.request.Request(
        url,
        data=json.dumps({"jsonrpc": "2.0", "id": 1, "method": "tools/call", "params": {"name": name, "arguments": arguments}}).encode(),
        headers={**HEADERS, "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            envelope = json.load(response)
        return json.loads(envelope["result"]["content"][0]["text"])
    except (OSError, urllib.error.URLError, json.JSONDecodeError, KeyError, IndexError, TypeError) as exc:
        raise RuntimeError(f"{name} returned an invalid response") from exc


def upload_xlsx() -> str:
    workbook = Workbook()
    sheet = workbook.create_sheet("Data")
    workbook.remove(workbook["Sheet"])
    if not isinstance(sheet, Worksheet):
        raise RuntimeError("expected a normal worksheet")
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Group"
    sheet.append(["value_a", "value_b"])
    sheet.append([1, 2])
    sheet.append([3, 4])
    buffer = io.BytesIO()
    workbook.save(buffer)
    request = urllib.request.Request(
        UPLOAD_URL,
        data=buffer.getvalue(),
        method="PUT",
        headers={**HEADERS, "X-Upload-Filename": urllib.parse.quote("merged.xlsx"), "X-Upload-Session-Id": SESSION_ID, "Content-Type": "application/octet-stream"},
    )
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            return str(json.load(response)["dataset_id"])
    except (OSError, urllib.error.URLError, json.JSONDecodeError, KeyError, TypeError) as exc:
        raise RuntimeError("XLSX upload failed") from exc


def delete_upload(dataset_id: str) -> None:
    request = urllib.request.Request(UPLOAD_URL + "/" + dataset_id, method="DELETE", headers={**HEADERS, "X-Upload-Session-Id": SESSION_ID})
    try:
        urllib.request.urlopen(request, timeout=30).close()
    except (OSError, urllib.error.URLError):
        pass


def main() -> int:
    if len(TOKEN) < 32:
        raise RuntimeError("MCP_SHARED_TOKEN is required")
    created: list[str] = []
    try:
        source_dataset_id = upload_xlsx()
        created.append(source_dataset_id)
        inspected = rpc(GRAFANA_QUERY_URL, "inspect_dataset", {"dataset_id": source_dataset_id})
        document_ref = inspected.get("document_ref")
        if not inspected.get("ok") or not isinstance(document_ref, str):
            raise RuntimeError(f"upload inspection did not return document_ref: {inspected}")

        preprocessing_code = """from openpyxl import load_workbook
workbook = load_workbook(document_path, data_only=True)
sheet = workbook['Data']
grid = [[sheet.cell(row, col).value for col in range(1, sheet.max_column + 1)] for row in range(1, sheet.max_row + 1)]
for merged in sheet.merged_cells.ranges:
    anchor = grid[merged.min_row - 1][merged.min_col - 1]
    for row in range(merged.min_row - 1, merged.max_row):
        for col in range(merged.min_col - 1, merged.max_col):
            grid[row][col] = anchor
columns = [f'{group}_{field}' for group, field in zip(grid[0], grid[1], strict=True)]
cleaned = pd.DataFrame(grid[2:], columns=columns)
emit_frame(cleaned, name='normalized-merged-headers')
emit({'columns': columns, 'rows': len(cleaned)}, name='result.json')
"""
        preprocessed = rpc(SANDBOX_URL, "execute_python_preprocessing", {"document_ref": document_ref, "python_code": preprocessing_code, "seed": 42})
        derived_frame_ref = preprocessed.get("derived_frame_ref")
        derived_dataset_id = preprocessed.get("derived_dataset_id")
        if not preprocessed.get("ok") or not isinstance(derived_frame_ref, str) or not isinstance(derived_dataset_id, str):
            raise RuntimeError(f"preprocessing did not produce both derived outputs: {preprocessed}")
        created.append(derived_dataset_id)

        chained = rpc(SANDBOX_URL, "execute_python_analysis", {"frame_ref": derived_frame_ref, "python_code": "emit({'rows': len(df), 'sum_a': int(df['Group_value_a'].sum())}, name='result.json')", "seed": 42})
        inline = chained.get("output_summary", {}).get("inline_results", [])
        if not chained.get("ok") or not inline or inline[0].get("value") != {"rows": 2, "sum_a": 4}:
            raise RuntimeError(f"derived frame was not reusable: {chained}")

        discovered = rpc(GRAFANA_QUERY_URL, "discover_datasets", {})
        derived_entry = next((item for item in discovered.get("datasets", []) if item.get("dataset_id") == derived_dataset_id), None)
        derived_inspection = rpc(GRAFANA_QUERY_URL, "inspect_dataset", {"dataset_id": derived_dataset_id})
        fields = [field["name"] for field in derived_inspection.get("metadata", {}).get("fields", [])]
        if derived_entry is None or fields != ["Group_value_a", "Group_value_b"]:
            raise RuntimeError("derived session dataset was not discoverable/inspectable")
        planned = rpc(PLANNER_URL, "plan_query", {"dataset_metadata_ref": derived_inspection["dataset_metadata_ref"], "selected_fields": fields, "minimum_rows": 1, "maximum_rows": 100})
        queried = rpc(GRAFANA_QUERY_URL, "execute_planned_query", {"plan_ref": planned.get("plan_ref"), "_server_session_id": SESSION_ID})
        if not planned.get("ok") or not queried.get("ok") or not isinstance(queried.get("frame_ref"), str):
            raise RuntimeError(f"derived session dataset was not reusable by Planner/Grafana Query: plan={planned}, query={queried}")

        print(json.dumps({"ok": True, "source_dataset_id": source_dataset_id, "derived_frame_ref": derived_frame_ref, "derived_dataset_id": derived_dataset_id, "query_frame_ref": queried["frame_ref"], "fields": fields, "chained_result": inline[0]["value"]}, ensure_ascii=False, indent=2))
        return 0
    finally:
        for dataset_id in reversed(created):
            delete_upload(dataset_id)


if __name__ == "__main__":
    raise SystemExit(main())
