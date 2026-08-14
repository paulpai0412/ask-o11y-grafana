#!/usr/bin/env python3
from __future__ import annotations

import io
import os
import tempfile
import zipfile
from pathlib import Path
import sys

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
import uploaded_datasets as uploads  # noqa: E402


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        uploads.UPLOAD_ROOT = Path(tmp)
        context = {"org_id": "1", "user_id": "7"}
        csv_result = uploads.store_upload(context=context, session_id="session-one", filename="sales.csv", raw=b"code,amount\n001,10\n002,20\n")
        assert csv_result["rows"] == 2 and csv_result["columns"] == 2
        assert csv_result["fields"] == [{"name": "code", "type": "string"}, {"name": "amount", "type": "number"}]
        wide_headers = [f"field_{index}" for index in range(uploads.MAX_UPLOAD_FIELDS)]
        wide_csv = (",".join(wide_headers) + "\n" + ",".join("1" for _ in wide_headers) + "\n").encode()
        assert uploads.store_upload(context=context, session_id="session-wide-csv", filename="wide.csv", raw=wide_csv)["columns"] == 200
        try:
            uploads.store_upload(context=context, session_id="session-too-wide", filename="too-wide.csv", raw=wide_csv.replace(b"\n", b",field_200\n", 1))
        except ValueError as exc:
            assert "exceeds 200 columns" in str(exc)
        else:
            raise AssertionError("201-column CSV upload was allowed")
        token = uploads.sign_csv_url(public_base="http://localhost:8772", secret="x" * 32, metadata=csv_result).rsplit("/", 1)[1]
        assert uploads.read_signed_csv(token, "x" * 32).read_text().startswith("code,amount")
        try:
            uploads.inspect_upload({"org_id": "1", "user_id": "8"}, csv_result["id"])
        except PermissionError:
            pass
        else:
            raise AssertionError("cross-user upload access was allowed")
        try:
            uploads.inspect_upload(context, csv_result["id"], "different-session")
        except PermissionError:
            pass
        else:
            raise AssertionError("cross-session upload access was allowed")
        assert uploads.list_uploads({**context, "session_id": "different-session"}) == []

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet("Data")
        if not isinstance(sheet, Worksheet):
            raise AssertionError("expected a normal worksheet")
        sheet.cell(row=1, column=1, value="date")
        sheet.cell(row=1, column=2, value="value")
        sheet.cell(row=2, column=1, value="2026-01-01")
        sheet.cell(row=2, column=2, value=3)
        for column in range(3, uploads.MAX_UPLOAD_FIELDS + 1):
            sheet.cell(row=1, column=column, value=f"field_{column}")
            sheet.cell(row=2, column=column, value=column)
        sheet.cell(row=3, column=1, value="formula")
        sheet.cell(row=3, column=2, value="=SUM(B2:B2)")
        workbook.create_sheet("Other")
        buffer = io.BytesIO()
        workbook.save(buffer)
        try:
            uploads.store_upload(context=context, session_id="session-two", filename="book.xlsx", raw=buffer.getvalue())
        except ValueError as exc:
            assert str(exc).startswith("SHEET_SELECTION_REQUIRED:")
        else:
            raise AssertionError("multi-sheet workbook did not require selection")
        xlsx_result = uploads.store_upload(context=context, session_id="session-two", filename="book.xlsx", raw=buffer.getvalue(), sheet="Data")
        assert xlsx_result["rows"] == 2 and xlsx_result["columns"] == 200 and xlsx_result["sheet"] == "Data"
        normalized = (uploads.UPLOAD_ROOT / xlsx_result["id"] / "data.csv").read_text()
        assert "=SUM" not in normalized

        for forbidden_entry in ("xl/vbaProject.bin", "xl/externalLinks/externalLink1.xml"):
            unsafe = io.BytesIO(buffer.getvalue())
            with zipfile.ZipFile(unsafe, "a") as archive:
                archive.writestr(forbidden_entry, b"unsafe")
            try:
                uploads.store_upload(context=context, session_id="session-three", filename="unsafe.xlsx", raw=unsafe.getvalue(), sheet="Data")
            except ValueError as exc:
                assert "not allowed" in str(exc)
            else:
                raise AssertionError(f"unsafe XLSX entry was allowed: {forbidden_entry}")

        expanded_limit = uploads.MAX_XLSX_EXPANDED_BYTES
        uploads.MAX_XLSX_EXPANDED_BYTES = 1
        try:
            uploads.store_upload(context=context, session_id="session-four", filename="bomb.xlsx", raw=buffer.getvalue(), sheet="Data")
        except ValueError as exc:
            assert "expanded size" in str(exc)
        else:
            raise AssertionError("XLSX expanded size limit was not enforced")
        finally:
            uploads.MAX_XLSX_EXPANDED_BYTES = expanded_limit

        try:
            removed = uploads.cleanup_expired(now=int(csv_result["expires_at"]) + 1)
        except (OSError, ValueError) as exc:
            raise AssertionError("TTL cleanup failed") from exc
        assert removed >= 1

        try:
            uploads.read_limited(io.BytesIO(b"x"), uploads.MAX_UPLOAD_BYTES + 1)
        except ValueError:
            pass
        else:
            raise AssertionError("oversized upload was allowed")
    print("uploaded dataset checks passed")
    return 0


if __name__ == "__main__":
    os.chdir(Path(__file__).resolve().parents[1])
    raise SystemExit(main())
